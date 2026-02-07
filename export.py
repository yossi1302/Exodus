"""
Export functionality for schedules to PDF and Excel formats
"""

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os


DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
TIMESLOTS = ['08:30', '11:00', '13:30', '16:00']


def export_to_pdf(schedule_data, program, schedule_id):
    """Export schedule to PDF format matching original layout"""
    
    filename = f'schedule_{program}_{schedule_id}.pdf'
    filepath = os.path.join('data', 'schedules', filename)
    
    doc = SimpleDocTemplate(filepath, pagesize=landscape(A4))
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#003366'),
        spaceAfter=12,
    )
    
    # Title
    metadata = schedule_data.get('metadata', {})
    title_text = f"Schedule {program} - {metadata.get('period', '')} {metadata.get('year', '')}"
    title = Paragraph(title_text, title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Process each week
    for week_key in sorted(schedule_data.get('schedule', {}).keys()):
        week_data = schedule_data['schedule'][week_key]
        week_num = week_key.replace('week_', '')
        
        # Week header
        week_title = Paragraph(f"<b>Week {week_num}</b>", styles['Heading2'])
        elements.append(week_title)
        elements.append(Spacer(1, 0.1*inch))
        
        # Build table data
        table_data = []
        
        # Header row
        header_row = ['Time'] + DAYS
        table_data.append(header_row)
        
        # Data rows
        for timeslot in TIMESLOTS:
            end_time = '10:30' if timeslot == '08:30' else ('13:00' if timeslot == '11:00' else ('15:30' if timeslot == '13:30' else '18:00'))
            row = [f"{timeslot} – {end_time}"]
            
            for day in DAYS:
                sessions = week_data.get(day, {}).get(timeslot, [])
                if sessions:
                    cell_content = []
                    for session in sessions:
                        course_abbr = session['course_name'].split('(')[1].replace(')', '') if '(' in session['course_name'] else session['course_name'][:10]
                        session_type = session['type'][0].upper() if session['type'] else ''
                        cell_text = f"{session['course']} ({session_type})\n{session['room']}"
                        cell_content.append(cell_text)
                    row.append('\n---\n'.join(cell_content))
                else:
                    row.append('—')
            
            table_data.append(row)
        
        # Create table
        table = Table(table_data, colWidths=[1.2*inch] + [1.5*inch]*5)
        
        # Style table
        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Time column
            ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#f8f9fa')),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (0, -1), 8),
            
            # Data cells
            ('FONTSIZE', (1, 1), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        elements.append(table)
        elements.append(PageBreak())
    
    # Build PDF
    doc.build(elements)
    
    return filepath


def export_to_excel(schedule_data, program, schedule_id):
    """Export schedule to Excel format"""
    
    filename = f'schedule_{program}_{schedule_id}.xlsx'
    filepath = os.path.join('data', 'schedules', filename)
    
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Styles
    header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    time_fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
    time_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    lecture_fill = PatternFill(start_color='e8f5e9', end_color='e8f5e9', fill_type='solid')
    tutorial_fill = PatternFill(start_color='fff3e0', end_color='fff3e0', fill_type='solid')
    lab_fill = PatternFill(start_color='f3e5f5', end_color='f3e5f5', fill_type='solid')
    
    # Create sheet for each week
    for week_key in sorted(schedule_data.get('schedule', {}).keys()):
        week_data = schedule_data['schedule'][week_key]
        week_num = week_key.replace('week_', '')
        
        ws = wb.create_sheet(title=f"Week {week_num}")
        
        # Set column widths
        ws.column_dimensions['A'].width = 18
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 25
        
        # Header
        ws['A1'] = 'Time'
        for col_idx, day in enumerate(DAYS, start=2):
            ws.cell(row=1, column=col_idx).value = day
        
        # Style header
        for col in range(1, 7):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        for row_idx, timeslot in enumerate(TIMESLOTS, start=2):
            end_time = '10:30' if timeslot == '08:30' else ('13:00' if timeslot == '11:00' else ('15:30' if timeslot == '13:30' else '18:00'))
            
            # Time column
            time_cell = ws.cell(row=row_idx, column=1)
            time_cell.value = f"{timeslot} – {end_time}"
            time_cell.fill = time_fill
            time_cell.font = time_font
            time_cell.alignment = Alignment(horizontal='center', vertical='center')
            time_cell.border = border
            
            # Day columns
            for col_idx, day in enumerate(DAYS, start=2):
                sessions = week_data.get(day, {}).get(timeslot, [])
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if sessions:
                    cell_lines = []
                    cell_fill = None
                    
                    for session in sessions:
                        course_abbr = session['course_name'].split('(')[1].replace(')', '') if '(' in session['course_name'] else session['course_name'][:10]
                        session_type = session['type'].capitalize()
                        line = f"{session['course']} ({session_type})\n{session['room']}"
                        cell_lines.append(line)
                        
                        # Set fill color based on session type
                        if session['type'] == 'lecture':
                            cell_fill = lecture_fill
                        elif session['type'] == 'tutorial':
                            cell_fill = tutorial_fill
                        elif session['type'] == 'lab':
                            cell_fill = lab_fill
                    
                    cell.value = '\n---\n'.join(cell_lines)
                    if cell_fill:
                        cell.fill = cell_fill
                else:
                    cell.value = '—'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.border = border
        
        # Set row heights
        for row in range(2, 6):
            ws.row_dimensions[row].height = 60
    
    # Save
    wb.save(filepath)
    
    return filepath


if __name__ == "__main__":
    # Test exports
    import json
    
    test_schedule = {
        "metadata": {"period": "Period 2", "year": "2024-2025"},
        "programs": {"CS_Y1": {"size": 300, "courses": ["BCS1220"]}},
        "schedule": {
            "week_1": {
                "Monday": {
                    "08:30": [{"course": "BCS1220", "course_name": "Objects in Programming (OIP)", 
                              "type": "lecture", "room": "MSP", "teacher": "E. Smirnov", "program": "CS_Y1"}],
                    "11:00": [],
                    "13:30": [],
                    "16:00": []
                },
                "Tuesday": {"08:30": [], "11:00": [], "13:30": [], "16:00": []},
                "Wednesday": {"08:30": [], "11:00": [], "13:30": [], "16:00": []},
                "Thursday": {"08:30": [], "11:00": [], "13:30": [], "16:00": []},
                "Friday": {"08:30": [], "11:00": [], "13:30": [], "16:00": []}
            }
        }
    }
    
    pdf_path = export_to_pdf(test_schedule, "CS_Y1", "test")
    excel_path = export_to_excel(test_schedule, "CS_Y1", "test")
    
    print(f"PDF created: {pdf_path}")
    print(f"Excel created: {excel_path}")
